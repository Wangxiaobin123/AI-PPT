"""Excel spreadsheet generator using openpyxl."""

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from .base import (
    BaseGenerator,
    GenerationMetadata,
    StructuredContent,
)


DEFAULT_COLORS = {
    "primary": "2B579A",
    "header_bg": "2B579A",
    "header_font": "FFFFFF",
    "alt_row": "F2F2F2",
    "border": "CCCCCC",
    "text": "333333",
}

DEFAULT_FONTS = {
    "header": "Calibri",
    "body": "Calibri",
}


def _resolve_colors(metadata: GenerationMetadata) -> dict[str, str]:
    colors = dict(DEFAULT_COLORS)
    colors.update(metadata.style.get("color_scheme", {}))
    return colors


def _resolve_fonts(metadata: GenerationMetadata) -> dict[str, str]:
    fonts = dict(DEFAULT_FONTS)
    fonts.update(metadata.style.get("fonts", {}))
    return fonts


def _auto_column_width(ws, col_idx: int, min_width: float = 8.0, max_width: float = 50.0):
    """Calculate and set column width based on cell contents."""
    col_letter = get_column_letter(col_idx)
    max_len = 0

    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
        for cell in row:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len

    # Add padding and clamp
    adjusted_width = max(min_width, min(max_len + 3, max_width))
    ws.column_dimensions[col_letter].width = adjusted_width


def _is_numeric(value: Any) -> bool:
    """Check if a value can be interpreted as a number."""
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        try:
            float(value.replace(",", ""))
            return True
        except (ValueError, AttributeError):
            return False
    return False


def _coerce_value(value: Any) -> Any:
    """Try to convert string values to appropriate Python types."""
    if not isinstance(value, str):
        return value

    stripped = value.strip()

    # Empty string
    if not stripped:
        return stripped

    # Percentage
    if stripped.endswith("%"):
        try:
            return float(stripped[:-1]) / 100
        except ValueError:
            return value

    # Number with commas (e.g., "1,234.56")
    try:
        return float(stripped.replace(",", ""))
    except ValueError:
        pass

    return value


class XLSXGenerator(BaseGenerator):
    """Generates Excel (.xlsx) spreadsheets from structured content."""

    async def generate(
        self, content: StructuredContent, metadata: GenerationMetadata
    ) -> bytes:
        """Generate an XLSX file and return it as bytes."""
        wb = Workbook()
        colors = _resolve_colors(metadata)
        fonts = _resolve_fonts(metadata)

        sheets_data = content.sheets
        if not sheets_data:
            # If no sheets defined, create a default sheet from sections
            sheets_data = self._sections_to_sheets(content)

        # Remove default empty sheet if we have explicit data
        if sheets_data:
            # Use the first sheet data for the default sheet
            first_sheet = sheets_data[0]
            ws = wb.active
            ws.title = first_sheet.get("name", "Sheet1")
            self._populate_sheet(ws, first_sheet, colors, fonts)

            # Additional sheets
            for sheet_data in sheets_data[1:]:
                ws = wb.create_sheet(title=sheet_data.get("name", "Sheet"))
                self._populate_sheet(ws, sheet_data, colors, fonts)
        else:
            # No data at all — create a titled sheet
            ws = wb.active
            ws.title = content.title or "Sheet1"
            ws["A1"] = content.title or "Empty Spreadsheet"
            ws["A1"].font = Font(
                name=fonts["header"], size=14, bold=True, color=colors["primary"]
            )

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------ #
    #  Sheet population
    # ------------------------------------------------------------------ #

    def _populate_sheet(
        self,
        ws,
        sheet_data: dict,
        colors: dict[str, str],
        fonts: dict[str, str],
    ):
        """Fill a worksheet with headers and row data."""
        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])
        number_formats = sheet_data.get("number_formats", {})
        # number_formats: {col_index: format_string}  e.g., {2: '#,##0.00', 3: '0.0%'}

        # Styles
        header_font = Font(
            name=fonts["header"],
            size=11,
            bold=True,
            color=colors.get("header_font", "FFFFFF"),
        )
        header_fill = PatternFill(
            start_color=colors.get("header_bg", "2B579A"),
            end_color=colors.get("header_bg", "2B579A"),
            fill_type="solid",
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        thin_border = Border(
            left=Side(style="thin", color=colors.get("border", "CCCCCC")),
            right=Side(style="thin", color=colors.get("border", "CCCCCC")),
            top=Side(style="thin", color=colors.get("border", "CCCCCC")),
            bottom=Side(style="thin", color=colors.get("border", "CCCCCC")),
        )
        body_font = Font(
            name=fonts["body"],
            size=10,
            color=colors.get("text", "333333"),
        )
        alt_fill = PatternFill(
            start_color=colors.get("alt_row", "F2F2F2"),
            end_color=colors.get("alt_row", "F2F2F2"),
            fill_type="solid",
        )
        body_alignment = Alignment(vertical="center", wrap_text=True)

        current_row = 1

        # Write header row
        if headers:
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=str(header))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            current_row += 1

            # Freeze the header row
            ws.freeze_panes = "A2"

        # Write data rows
        for row_offset, row_data in enumerate(rows):
            for col_idx, value in enumerate(row_data, start=1):
                coerced = _coerce_value(value)
                cell = ws.cell(row=current_row, column=col_idx, value=coerced)
                cell.font = body_font
                cell.border = thin_border
                cell.alignment = body_alignment

                # Alternating row fill
                if row_offset % 2 == 1:
                    cell.fill = alt_fill

                # Apply number format if specified
                col_key = col_idx - 1  # 0-based index
                if col_key in number_formats:
                    cell.number_format = number_formats[col_key]
                elif isinstance(coerced, float):
                    # Check if it was originally a percentage
                    orig = str(value).strip() if isinstance(value, str) else ""
                    if orig.endswith("%"):
                        cell.number_format = "0.0%"
                    else:
                        cell.number_format = "#,##0.00"

            current_row += 1

        # Auto-size columns
        num_cols = max(
            len(headers),
            max((len(r) for r in rows), default=0),
        ) if (headers or rows) else 0

        for col_idx in range(1, num_cols + 1):
            _auto_column_width(ws, col_idx)

        # Add auto-filter on header row if headers exist
        if headers:
            last_col_letter = get_column_letter(len(headers))
            last_row = 1 + len(rows)
            ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # ------------------------------------------------------------------ #
    #  Helpers
    # ------------------------------------------------------------------ #

    @staticmethod
    def _sections_to_sheets(content: StructuredContent) -> list[dict]:
        """Convert document sections to sheet data as a fallback."""
        sheets = []
        for section in content.sections:
            for block in section.blocks:
                if block.type.value == "table" and block.rows:
                    headers = [str(c) for c in block.rows[0]] if block.rows else []
                    data_rows = block.rows[1:] if len(block.rows) > 1 else []
                    sheets.append({
                        "name": section.title[:31] if section.title else "Sheet1",
                        "headers": headers,
                        "rows": data_rows,
                    })
        return sheets
