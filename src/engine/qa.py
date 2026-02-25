"""Quality-assurance validator for skill outputs.

Runs a battery of generic and format-specific checks against a
:class:`SkillResult` and produces a :class:`QAReport`.
"""

from __future__ import annotations

import io
from typing import Callable

from src.skills.models import QACheck, QAReport, SkillResult
from src.utils.logging import get_logger

logger = get_logger("engine.qa")

# ---------------------------------------------------------------------------
# Size thresholds (bytes)
# ---------------------------------------------------------------------------
MIN_OUTPUT_BYTES = 100  # anything smaller is almost certainly an error
MAX_OUTPUT_BYTES = 100 * 1024 * 1024  # 100 MiB upper bound


class QAValidator:
    """Stateless validator that inspects a :class:`SkillResult`.

    Checks performed:

    1. **non_empty** -- content_bytes is not empty.
    2. **reasonable_size** -- content_bytes falls within acceptable bounds.
    3. **format_specific** -- deeper inspection based on ``output_format``:
       - *pptx*: open with ``python-pptx`` and verify at least one slide.
       - *docx*: open with ``python-docx`` and verify at least one paragraph.
       - *xlsx*: open with ``openpyxl`` and verify at least one sheet.
       - *pdf*: verify the PDF header bytes (``%PDF``).
       - *html*: verify the presence of an ``<html`` tag.
    """

    def _get_format_checkers(self) -> dict[str, Callable[..., QACheck]]:
        """Return a mapping of output_format -> checker method."""
        return {
            "pptx": self._check_pptx,
            "docx": self._check_docx,
            "xlsx": self._check_xlsx,
            "pdf": self._check_pdf,
            "html": self._check_html,
        }

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    async def validate(self, skill_result: SkillResult) -> QAReport:
        """Run all applicable checks and return a :class:`QAReport`."""
        checks: list[QACheck] = []

        # 1. Non-empty check
        checks.append(self._check_non_empty(skill_result))

        # 2. Reasonable size check
        checks.append(self._check_reasonable_size(skill_result))

        # 3. Format-specific checks (only when there are content bytes)
        if skill_result.content_bytes:
            fmt = skill_result.output_format.lower().strip()
            format_checkers = self._get_format_checkers()
            format_checker = format_checkers.get(fmt)
            if format_checker is not None:
                check = format_checker(skill_result)
                checks.append(check)
            else:
                logger.debug("no_format_checker", output_format=fmt)

        all_passed = all(c.passed for c in checks)

        report = QAReport(checks=checks, passed=all_passed)
        logger.info(
            "qa_validation_complete",
            passed=all_passed,
            total_checks=len(checks),
            failed_checks=[c.name for c in checks if not c.passed],
        )
        return report

    # ------------------------------------------------------------------
    # Generic checks
    # ------------------------------------------------------------------

    @staticmethod
    def _check_non_empty(result: SkillResult) -> QACheck:
        """Verify that the skill produced some output bytes."""
        has_content = len(result.content_bytes) > 0
        return QACheck(
            name="non_empty",
            passed=has_content,
            detail=(
                f"Size: {len(result.content_bytes)} bytes"
                if has_content
                else "Skill produced zero bytes of output"
            ),
        )

    @staticmethod
    def _check_reasonable_size(result: SkillResult) -> QACheck:
        """Verify that the output size falls within acceptable bounds."""
        size = len(result.content_bytes)
        if size == 0:
            # Already covered by non_empty; pass here to avoid double penalty.
            return QACheck(
                name="reasonable_size",
                passed=True,
                detail="Skipped (empty content handled by non_empty check)",
            )
        if size < MIN_OUTPUT_BYTES:
            return QACheck(
                name="reasonable_size",
                passed=False,
                detail=f"Output too small: {size} bytes (minimum: {MIN_OUTPUT_BYTES})",
            )
        if size > MAX_OUTPUT_BYTES:
            return QACheck(
                name="reasonable_size",
                passed=False,
                detail=f"Output too large: {size} bytes (maximum: {MAX_OUTPUT_BYTES})",
            )
        return QACheck(
            name="reasonable_size",
            passed=True,
            detail=f"Size OK: {size} bytes",
        )

    # ------------------------------------------------------------------
    # Format-specific checks
    # ------------------------------------------------------------------

    def _check_pptx(self, result: SkillResult) -> QACheck:
        """Open the PPTX and confirm at least one slide exists."""
        try:
            from pptx import Presentation  # type: ignore[import-untyped]

            prs = Presentation(io.BytesIO(result.content_bytes))
            slide_count = len(prs.slides)
            if slide_count == 0:
                return QACheck(
                    name="format_pptx",
                    passed=False,
                    detail="PPTX file contains zero slides",
                )
            return QACheck(
                name="format_pptx",
                passed=True,
                detail=f"PPTX valid with {slide_count} slide(s)",
            )
        except Exception as exc:
            return QACheck(
                name="format_pptx",
                passed=False,
                detail=f"Failed to open PPTX: {exc}",
            )

    def _check_docx(self, result: SkillResult) -> QACheck:
        """Open the DOCX and confirm at least one paragraph exists."""
        try:
            from docx import Document  # type: ignore[import-untyped]

            doc = Document(io.BytesIO(result.content_bytes))
            para_count = len(doc.paragraphs)
            if para_count == 0:
                return QACheck(
                    name="format_docx",
                    passed=False,
                    detail="DOCX file contains zero paragraphs",
                )
            return QACheck(
                name="format_docx",
                passed=True,
                detail=f"DOCX valid with {para_count} paragraph(s)",
            )
        except Exception as exc:
            return QACheck(
                name="format_docx",
                passed=False,
                detail=f"Failed to open DOCX: {exc}",
            )

    def _check_xlsx(self, result: SkillResult) -> QACheck:
        """Open the XLSX and confirm at least one sheet exists."""
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]

            wb = load_workbook(io.BytesIO(result.content_bytes), read_only=True)
            sheet_count = len(wb.sheetnames)
            wb.close()
            if sheet_count == 0:
                return QACheck(
                    name="format_xlsx",
                    passed=False,
                    detail="XLSX file contains zero sheets",
                )
            return QACheck(
                name="format_xlsx",
                passed=True,
                detail=f"XLSX valid with {sheet_count} sheet(s)",
            )
        except Exception as exc:
            return QACheck(
                name="format_xlsx",
                passed=False,
                detail=f"Failed to open XLSX: {exc}",
            )

    def _check_pdf(self, result: SkillResult) -> QACheck:
        """Verify that the content starts with the PDF magic bytes."""
        header = result.content_bytes[:1024]
        if header.startswith(b"%PDF"):
            return QACheck(
                name="format_pdf",
                passed=True,
                detail="PDF header bytes verified",
            )
        return QACheck(
            name="format_pdf",
            passed=False,
            detail="Content does not start with %PDF header",
        )

    def _check_html(self, result: SkillResult) -> QACheck:
        """Verify that the content contains an ``<html`` tag."""
        try:
            text = result.content_bytes.decode("utf-8", errors="replace").lower()
        except Exception:
            text = ""
        if "<html" in text:
            return QACheck(
                name="format_html",
                passed=True,
                detail="<html> tag found",
            )
        return QACheck(
            name="format_html",
            passed=False,
            detail="No <html> tag found in content",
        )
