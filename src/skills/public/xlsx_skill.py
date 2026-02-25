"""XLSX generation skill -- converts various input formats into Excel spreadsheets."""

from __future__ import annotations

import json
import traceback

from src.generators.base import (
    ContentBlock,
    ContentType,
    GenerationMetadata,
    StructuredContent,
)
from src.skills.base import BaseSkill
from src.skills.models import QACheck, QAReport, SkillMetadata, SkillResult
from src.utils.logging import get_logger

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# Content format detection
# ---------------------------------------------------------------------------

def _detect_format(content: str) -> str:
    """Heuristically detect the format of *content*."""
    stripped = content.strip()

    if stripped.startswith("{") or stripped.startswith("["):
        try:
            json.loads(stripped)
            return "json"
        except (json.JSONDecodeError, ValueError):
            pass

    first_line = stripped.split("\n", 1)[0]
    if "," in first_line and "<" not in first_line and first_line.count(",") >= 2:
        return "csv"

    if "<html" in stripped.lower() or stripped.startswith("<!") or "</" in stripped[:200]:
        return "html"

    if any(stripped.startswith(c) for c in ("# ", "## ", "### ")):
        return "markdown"

    return "text"


# ---------------------------------------------------------------------------
# Parsers -> StructuredContent with sheets
# ---------------------------------------------------------------------------

def _parse_csv_to_sheets(content: str, title: str) -> StructuredContent:
    """Parse CSV content into sheets data."""
    rows: list[list[str]] = []
    for line in content.strip().split("\n"):
        if line.strip():
            rows.append([cell.strip() for cell in line.split(",")])

    headers = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []

    sheet = {
        "name": title or "Sheet1",
        "headers": headers,
        "rows": data_rows,
    }
    return StructuredContent(title=title or "Data", sheets=[sheet])


def _parse_json_to_sheets(content: str, title: str) -> StructuredContent:
    """Parse JSON content into sheets."""
    data = json.loads(content)

    # Already shaped like StructuredContent with sheets.
    if isinstance(data, dict) and "sheets" in data:
        sc = StructuredContent.model_validate(data)
        if title:
            sc.title = title
        return sc

    # List of objects -> single sheet.
    if isinstance(data, list) and data and isinstance(data[0], dict):
        headers = list(data[0].keys())
        rows = [[str(item.get(h, "")) for h in headers] for item in data]
        sheet = {"name": title or "Sheet1", "headers": headers, "rows": rows}
        return StructuredContent(title=title or "Data", sheets=[sheet])

    # List of lists -> single sheet (first row = headers).
    if isinstance(data, list) and data and isinstance(data[0], list):
        headers = [str(h) for h in data[0]]
        rows = [[str(c) for c in row] for row in data[1:]]
        sheet = {"name": title or "Sheet1", "headers": headers, "rows": rows}
        return StructuredContent(title=title or "Data", sheets=[sheet])

    # Single object -> key-value sheet.
    if isinstance(data, dict):
        headers = ["Key", "Value"]
        rows = [[str(k), str(v)] for k, v in data.items()]
        sheet = {"name": title or "Sheet1", "headers": headers, "rows": rows}
        return StructuredContent(title=title or "Data", sheets=[sheet])

    # Fallback
    return _parse_text_to_sheets(json.dumps(data, indent=2), title)


def _parse_markdown_to_sheets(content: str, title: str) -> StructuredContent:
    """Extract tables from Markdown or fall back to line-based rows."""
    # Look for pipe-delimited tables.
    lines = content.strip().split("\n")
    table_rows: list[list[str]] = []
    in_table = False

    for line in lines:
        stripped = line.strip()
        if "|" in stripped:
            # Skip separator lines (e.g. |---|---|).
            if set(stripped.replace("|", "").replace("-", "").replace(" ", "")) == set():
                continue
            cells = [c.strip() for c in stripped.strip("|").split("|")]
            table_rows.append(cells)
            in_table = True
        elif in_table:
            break  # End of table block.

    if table_rows:
        headers = table_rows[0]
        data_rows = table_rows[1:]
        sheet = {"name": title or "Sheet1", "headers": headers, "rows": data_rows}
        return StructuredContent(title=title or "Data", sheets=[sheet])

    # No table found: each line becomes a row with a single column.
    return _parse_text_to_sheets(content, title)


def _parse_text_to_sheets(content: str, title: str) -> StructuredContent:
    """Plain text -> single-column sheet, one row per non-empty line."""
    lines = [line.strip() for line in content.split("\n") if line.strip()]
    headers = ["Content"]
    rows = [[line] for line in lines]
    sheet = {"name": title or "Sheet1", "headers": headers, "rows": rows}
    return StructuredContent(title=title or "Data", sheets=[sheet])


def _parse_html_to_sheets(content: str, title: str) -> StructuredContent:
    """Strip tags, then treat as text."""
    import re

    text = re.sub(r"<[^>]+>", " ", content)
    text = re.sub(r"\s+", " ", text).strip()
    return _parse_text_to_sheets(text, title)


# ---------------------------------------------------------------------------
# The skill
# ---------------------------------------------------------------------------

class XlsxSkill(BaseSkill):
    """Skill that generates XLSX (Excel) spreadsheets."""

    @property
    def metadata(self) -> SkillMetadata:
        return SkillMetadata(
            name="xlsx",
            description="Generate Excel spreadsheets from various input formats",
            input_formats=["html", "markdown", "json", "text", "csv"],
            output_format="xlsx",
            version="1.0.0",
            tags=["spreadsheet", "office", "xlsx"],
        )

    async def validate_params(self, params: dict) -> bool:
        if "content" not in params:
            return False
        if not isinstance(params["content"], str):
            return False
        if not params["content"].strip():
            return False
        return True

    async def execute(self, params: dict, context: dict) -> SkillResult:
        try:
            content_str: str = params["content"]
            content_format: str = params.get("content_format", "") or _detect_format(content_str)
            title: str = params.get("title", "")
            template: str = params.get("template", "default")
            style: dict = params.get("style", {})

            parser_map = {
                "csv": _parse_csv_to_sheets,
                "json": _parse_json_to_sheets,
                "markdown": _parse_markdown_to_sheets,
                "text": _parse_text_to_sheets,
                "html": _parse_html_to_sheets,
            }
            parser = parser_map.get(content_format, _parse_text_to_sheets)
            structured = parser(content_str, title)

            gen_meta = GenerationMetadata(template=template, style=style)

            try:
                from src.generators.xlsx_generator import XlsxGenerator
                generator = XlsxGenerator()
            except ImportError:
                logger.warning("xlsx_generator_not_found", detail="Using fallback generator")
                generator = _FallbackXlsxGenerator()

            output_bytes = await generator.generate(structured, gen_meta)

            return SkillResult(
                success=True,
                output_format="xlsx",
                content_bytes=output_bytes,
                metadata={
                    "title": structured.title,
                    "sheet_count": len(structured.sheets),
                    "content_format": content_format,
                    "template": template,
                },
            )

        except Exception as exc:
            logger.exception("xlsx_skill_error")
            return SkillResult(
                success=False,
                output_format="xlsx",
                error=f"{type(exc).__name__}: {exc}",
                metadata={"traceback": traceback.format_exc()},
            )

    async def qa_check(self, result: SkillResult) -> QAReport:
        report = await super().qa_check(result)
        if result.content_bytes:
            is_zip = result.content_bytes[:4] == b"PK\x03\x04"
            report.checks.append(
                QACheck(
                    name="valid_xlsx_header",
                    passed=is_zip,
                    detail="File starts with PK zip header" if is_zip else "Invalid XLSX header",
                )
            )
            report.passed = report.passed and is_zip
        return report


# ---------------------------------------------------------------------------
# Fallback generator
# ---------------------------------------------------------------------------

class _FallbackXlsxGenerator:
    """Minimal XLSX generator using openpyxl directly."""

    async def generate(self, content: StructuredContent, metadata: GenerationMetadata) -> bytes:
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        wb = Workbook()
        # Remove the default sheet; we create our own below.
        default_sheet = wb.active

        sheets_data = content.sheets if content.sheets else [
            {"name": "Sheet1", "headers": ["Content"], "rows": [["(empty)"]]}
        ]

        for idx, sheet_info in enumerate(sheets_data):
            if idx == 0 and default_sheet is not None:
                ws = default_sheet
                ws.title = sheet_info.get("name", "Sheet1")
            else:
                ws = wb.create_sheet(title=sheet_info.get("name", f"Sheet{idx + 1}"))

            headers = sheet_info.get("headers", [])
            rows = sheet_info.get("rows", [])

            # Write header row.
            header_font = Font(bold=True)
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Write data rows.
            for row_idx, row in enumerate(rows, start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-size columns (approximate).
            for col_idx in range(1, len(headers) + 1):
                max_len = max(
                    (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(rows) + 2)),
                    default=8,
                )
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
